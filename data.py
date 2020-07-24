"""
    番剧数据统计 (data.py) 及数据绘制 (analyze.py)
    1.url 通过 F12 --> Network --> XHR --> stat?season_id=xxx 获取
    2.sleep time 可自定义
    3.使用挂在服务器端运行 数据保存为 txt 和 excel 两种格式方便基础用户使用
    4.客户端可以使用 analyze.py 自动通过 SFTP 下载表格使用 matplotlib 绘制图像

                                        --------- Mr.Porridge 2020.02.14
"""

import requests
import time
import openpyxl


def write_excel_xlsx(path: str, sheet_name: str, dic: dict):
    book = openpyxl.load_workbook(path)
    # 取第一张表
    table = book.get_sheet_by_name(sheet_name)
    table = book.active
    # print(table.title)  # 输出表名
    rows = table.max_row  # 获得行数
    columns = table.max_column  # 获得列数
    values = [str(rows)]
    for item in dic.values():
        values.append(str(item))
    table.cell(rows + 1, 1).value = values[0]
    for col in range(2, columns + 1):
        table.cell(rows + 1, col).value = values[col - 1]
    book.save(path)

    data_text = open("log.txt", "a+")
    data_text.write(time.strftime("%Y-%m-%d-%H:%M:%S", time.localtime()) + "Data collected successfully！\n")
    data_text.close()


def get_api(api_url):
    headers = {
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8",
        "Accept-Encoding": "gzip, deflate, br",
        "Accept-Language": "zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2",
        "Host": "api.bilibili.com",
        "Referer": "https://www.bilibili.com/",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.77 Safari/537.36",
    }
    res = requests.get(api_url, headers=headers)
    res_dict = res.json()
    print(res_dict["result"])
    res = res_dict["result"]
    data = {
        "views": res["views"],
        "danmu": res["danmakus"],
        "coins": res["coins"],
        "series_follow": res["series_follow"],
        "time": time.strftime("%Y-%m-%d-%H:%M:%S", time.localtime()),
    }
    # 写入文本文件
    data_text = open("LEVEL5.txt", "a+")
    data_text.write(str(data) + "\n")
    data_text.close()
    # 写入Excel文件
    write_excel_xlsx("data.xlsx", "Sheet1", data)
    return res_dict


# F12 --> Network --> XHR --> stat?season_id=29325
url = "https://api.bilibili.com/pgc/web/season/stat?season_id=29325"
# sleep time
sleep_beauty = 3600

while True:
    get_api(url)
    time.sleep(sleep_beauty)
